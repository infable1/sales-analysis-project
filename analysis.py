import psycopg2
from config import host, user, password, db_name
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def analysis(table_name='sales'):
    try:
        connection = psycopg2.connect(
            host=host,
            user=user,
            password=password,
            database=db_name
        )
            
        connection.autocommit = True
        
        with connection.cursor() as cursor:
            wb = Workbook()

            default_sheet = "Sheet"
            if default_sheet in wb.sheetnames:
                sheet_to_remove = wb[default_sheet]
                wb.remove(sheet_to_remove)

            wb.create_sheet('Data')

            data_sheet = wb['Data']
            data_sheet['A1'].value = 'Total amount'
            data_sheet['A1'].font = Font(bold=True)
            data_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cursor.execute(f"SELECT SUM(price) AS total_sales FROM {table_name};")
            total_amount = cursor.fetchone()
            data_sheet['B1'].value = total_amount[0]
            data_sheet['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            print(f"Общая сумма продаж: {total_amount[0]}")
            
            data_sheet['A3'].value = 'Product'
            data_sheet['A3'].font = Font(bold=True)
            data_sheet['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_sheet['B3'].value = 'Average price'
            data_sheet['B3'].font = Font(bold=True)
            data_sheet['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cursor.execute(f"SELECT product, AVG(price) AS avg_price FROM {table_name} GROUP BY product;")
            results = cursor.fetchall()

            print("Средняя цена по продуктам:")
            i = 1
            for row in results:
                data_sheet[f'A{3 + i}'].value = row[0]
                data_sheet[f'A{3 + i}'].font = Font(bold=True)
                data_sheet[f'A{3 + i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                data_sheet[f'B{3 + i}'].value = row[1]
                data_sheet[f'B{3 + i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                i += 1
            
            data_sheet['D3'].value = 'Date'
            data_sheet['D3'].font = Font(bold=True)
            data_sheet['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_sheet['E3'].value = 'Sales amount'
            data_sheet['E3'].font = Font(bold=True)
            data_sheet['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cursor.execute(f"SELECT sale_date, SUM(price) AS daily_sales FROM {table_name} GROUP BY sale_date;")
            results = cursor.fetchall()

            i = 1
            for row in results:
                data_sheet[f'D{3 + i}'].value = row[0]
                data_sheet[f'D{3 + i}'].font = Font(bold=True)
                data_sheet[f'D{3 + i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                data_sheet[f'D{3 + i}'].number_format
                
                data_sheet[f'E{3 + i}'].value = row[1]
                data_sheet[f'E{3 + i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                i += 1
            
            wb.save('sales-analysis-project/data/data.xlsx')
            
    except Exception as _ex:
        print("[INFO] Error while working with PostgreSQL", _ex)
    finally:
        if connection:
            connection.close()